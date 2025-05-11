# 4. 整合所有功能的主程式
# 這個程式將整合前面三個程式的功能，提供一個完整的工作流程

import os
import sys
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 有條件地導入Windows特定模組
if sys.platform == 'win32':
    import win32com.client
    import pythoncom
def has_images(doc_path):
    """檢查Word文件中是否含有圖片"""
    try:
        doc = Document(doc_path)

        # 檢查是否有圖片
        for rel in doc.part.rels.values():
            if "image" in rel.reltype:
                return True

        return False
    except Exception as e:
        print(f"處理檔案 {doc_path} 時發生錯誤: {e}")
        return False

def analyze_folder(folder_path):
    """分析資料夾中的文件，識別問題文件"""
    # 存儲含有圖片的文件和非Word檔案
    image_files = []
    non_word_files = []
    all_files = []
    
    print(f"\n正在分析資料夾: {folder_path}")
    print("="*50)
    
    # 遍歷資料夾中的所有文件
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            file_path = os.path.join(root, filename)
            all_files.append(file_path)
            
            # 使用try-except來處理各種可能的錯誤情況
            try:
                if filename.endswith(".docx") or filename.endswith(".doc"):
                    if has_images(file_path):
                        image_files.append(file_path)
                        print(f"⚠️  {filename} - Word檔含有圖片")
                else:
                    if not filename.endswith(".pdf") and not filename.endswith(".txt"):
                        non_word_files.append(file_path)
                        print(f"⚠️  {filename} - 非Word檔案格式")
            except Exception as e:
                print(f"❌ 處理 {file_path} 時出錯: {e}")
    
    # 顯示詳細結果
    print("\n=== 分析結果摘要 ===")
    print(f"發現 {len(image_files)} 個含有圖片的Word文件")
    print(f"發現 {len(non_word_files)} 個非Word/PDF/TXT檔案")
    
    return {
        "image_files": image_files,
        "non_word_files": non_word_files,
        "all_files": all_files
    }

def convert_word_to_pdf(word_files, output_folder):
    """將Word文件轉換為PDF"""
    # 創建目標資料夾
    os.makedirs(output_folder, exist_ok=True)
    
    # 計數器
    converted_count = 0
    error_count = 0
    
    print(f"\n正在將Word文件轉換為PDF")
    print(f"輸出資料夾: {output_folder}")
    print("="*50)
    
    for word_path in word_files:
        if word_path.endswith(".docx") or word_path.endswith(".doc"):
            filename = os.path.basename(word_path)
            pdf_filename = os.path.splitext(filename)[0] + ".pdf"
            pdf_path = os.path.join(output_folder, pdf_filename)
            
            print(f"正在轉換: {filename} -> {pdf_filename}")
            
            try:
                # 初始化COM
                pythoncom.CoInitialize()
                
                # 創建Word應用實例
                word = win32com.client.Dispatch("Word.Application")
                word.Visible = False
                
                # 打開Word文件
                doc = word.Documents.Open(word_path)
                
                # 保存為PDF
                doc.SaveAs(pdf_path, FileFormat=17)  # 17代表PDF格式
                
                # 關閉文件和應用
                doc.Close()
                word.Quit()
                
                print(f"✅ 成功轉換: {filename}")
                converted_count += 1
                
            except Exception as e:
                print(f"❌ 轉換失敗: {filename}, 錯誤: {e}")
                error_count += 1
                
            finally:
                # 釋放COM資源
                pythoncom.CoUninitialize()
    
    # 顯示詳細結果
    print("\n=== 轉換結果摘要 ===")
    print(f"共處理 {converted_count + error_count} 個Word文件")
    print(f"成功轉換: {converted_count} 個")
    print(f"轉換失敗: {error_count} 個")
    
    return converted_count, error_count

def convert_word_to_pdf_macos(word_files, output_folder):
    """使用 macOS 方法將 Word 文件轉換為 PDF"""
    os.makedirs(output_folder, exist_ok=True)
    
    converted_count = 0
    error_count = 0
    
    print(f"\n正在將Word文件轉換為PDF")
    print(f"輸出資料夾: {output_folder}")
    print("="*50)
    
    # 先嘗試退出任何可能已經在運行的Word實例
    try:
        quit_script = """
        tell application "Microsoft Word"
            if it is running then
                quit
            end if
        end tell
        """
        subprocess.run(["osascript", "-e", quit_script], capture_output=True, text=True)
        import time
        time.sleep(2)  # 等待Word完全關閉
    except:
        pass
    
    for word_path in word_files:
        if word_path.endswith(".docx") or word_path.endswith(".doc"):
            filename = os.path.basename(word_path)
            pdf_filename = os.path.splitext(filename)[0] + ".pdf"
            pdf_path = os.path.join(output_folder, pdf_filename)
            
            print(f"正在轉換: {filename} -> {pdf_filename}")
            
            try:
                # 使用修改後的AppleScript，讓Word有更多時間處理文件
                script = f"""
                tell application "Microsoft Word"
                    activate
                    delay 2
                    open "{word_path}"
                    delay 3
                    set theDoc to active document
                    save as theDoc file name "{pdf_path}" file format format PDF
                    delay 2
                    close theDoc saving no
                    delay 1
                end tell
                """
                
                # 執行AppleScript
                import subprocess
                result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
                
                if result.returncode == 0:
                    print(f"✅ 成功轉換: {filename}")
                    converted_count += 1
                else:
                    print(f"❌ 轉換失敗: {filename}, 錯誤: {result.stderr}")
                    
                    # 嘗試替代方法 - 使用automator工作流程或系統PDF列印功能
                    print(f"   嘗試替代方法...")
                    try:
                        # 使用另一種方法
                        alt_script = f"""
                        tell application "Microsoft Word"
                            activate
                            delay 2
                            open "{word_path}"
                            delay 3
                            set myDoc to active document
                            set myFilePath to "{pdf_path}"
                            make new PDF file at active document with properties {{file name:myFilePath}}
                            delay 2
                            close active document saving no
                            delay 1
                        end tell
                        """
                        subprocess.run(["osascript", "-e", alt_script], check=True)
                        print(f"✅ 使用替代方法成功轉換: {filename}")
                        converted_count += 1
                    except Exception as e:
                        print(f"❌ 替代方法也失敗: {filename}, 錯誤: {e}")
                        error_count += 1
                
            except Exception as e:
                print(f"❌ 轉換失敗: {filename}, 錯誤: {e}")
                error_count += 1
            
            # 每處理5個文件後，嘗試退出並重新啟動Word以釋放內存
            if (converted_count + error_count) % 5 == 0:
                try:
                    quit_restart_script = """
                    tell application "Microsoft Word"
                        quit
                    end tell
                    """
                    subprocess.run(["osascript", "-e", quit_restart_script], capture_output=True, text=True)
                    import time
                    time.sleep(3)  # 給Word足夠時間關閉
                except:
                    pass
    
    # 確保在完成後關閉Word
    try:
        final_quit_script = """
        tell application "Microsoft Word"
            quit
        end tell
        """
        subprocess.run(["osascript", "-e", final_quit_script], capture_output=True, text=True)
    except:
        pass
    
    # 顯示詳細結果
    print("\n=== 轉換結果摘要 ===")
    print(f"共處理 {converted_count + error_count} 個Word文件")
    print(f"成功轉換: {converted_count} 個")
    print(f"轉換失敗: {error_count} 個")
    
    return converted_count, error_count

def generate_excel_report(image_files, non_word_files):
    """生成Excel報告"""
    today = datetime.datetime.now().strftime("%Y%m%d")
    excel_path = os.path.expanduser(f"~/Documents/會資/Final Project/{today} Doubtful File Name.xlsx") 
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    
    # 創建DataFrame
    data = []
    
    # 添加含有圖片的Word文件
    for file_path in image_files:
        filename = os.path.basename(file_path)
        data.append({
            "名稱": filename,
            "疑慮原因": "含有圖片",
            "完成核取方塊": False
        })
    
    # 添加非Word檔案（排除.DS_Store文件）
    for file_path in non_word_files:
        filename = os.path.basename(file_path)
        # 跳過.DS_Store文件
        if filename == ".DS_Store":
            continue
        data.append({
            "疑慮原因": "疑慮性合約",
            "名稱": filename,
            "完成核取方塊": False
        })
    
    # 創建DataFrame
    if data:
        df = pd.DataFrame(data)
        
        # 交換列順序以符合要求
        df = df[["名稱", "疑慮原因", "完成核取方塊"]]
        
        # 創建Excel工作簿
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='疑慮文件')
        
        # 獲取工作表進行格式設置
        workbook = writer.book
        worksheet = writer.sheets['疑慮文件']
        
        # 設置表頭格式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 設置邊框
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 應用表頭格式
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 應用數據格式
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, 4):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx == 3:  # 完成核取方塊列
                    cell.alignment = Alignment(horizontal='center')
        
        # 調整列寬 - 讓所有列都有足夠寬度
        for col_idx, column in enumerate(df.columns, 1):
            # 針對特定列設置最小寬度
            if column == "名稱":
                min_width = 40  # 設置名稱列的最小寬度
            elif column == "疑慮原因":
                min_width = 30  # 設置疑慮原因列的最小寬度
            elif column == "完成核取方塊":
                min_width = 20  # 設置完成核取方塊列的最小寬度
            else:
                min_width = 15  # 其他列的最小寬度
            
            # 計算實際需要的寬度（基於內容長度）
            content_width = max(len(column) * 1.2, df[column].astype(str).map(len).max() * 1.2)
            
            # 使用兩者中較大的值
            column_width = max(min_width, content_width)
            
            # 設置列寬
            worksheet.column_dimensions[get_column_letter(col_idx)].width = column_width
        
        # 保存Excel文件
        writer.close()
        
        print(f"\n報告已生成: {excel_path}")
        return excel_path, len(data)
    else:
        print("\n沒有發現問題文件，無需生成報告")
        return None, 0

def analyze_folder(folder_path):
    """分析資料夾中的文件，識別問題文件"""
    # 存儲含有圖片的文件和非Word檔案
    image_files = []
    non_word_files = []
    all_files = []
    
    print(f"\n正在分析資料夾: {folder_path}")
    print("="*50)
    
    # 遍歷資料夾中的所有文件
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            # 跳過.DS_Store文件
            if filename == ".DS_Store":
                continue
                
            file_path = os.path.join(root, filename)
            all_files.append(file_path)
            
            # 使用try-except來處理各種可能的錯誤情況
            try:
                if filename.endswith(".docx") or filename.endswith(".doc"):
                    if has_images(file_path):
                        image_files.append(file_path)
                        print(f"⚠️  {filename} - Word檔含有圖片")
                else:
                    if not filename.endswith(".pdf") and not filename.endswith(".txt"):
                        non_word_files.append(file_path)
                        print(f"⚠️  {filename} - 非Word檔案格式")
            except Exception as e:
                print(f"❌ 處理 {file_path} 時出錯: {e}")
    
    # 顯示詳細結果
    print("\n=== 分析結果摘要 ===")
    print(f"發現 {len(image_files)} 個含有圖片的Word文件")
    print(f"發現 {len(non_word_files)} 個非Word/PDF/TXT檔案")
    
    return {
        "image_files": image_files,
        "non_word_files": non_word_files,
        "all_files": all_files
    }

def select_folder_with_gui():
    """使用對話框選擇資料夾"""
    root = tk.Tk()
    root.withdraw()  # 隱藏主窗口
    
    folder_path = filedialog.askdirectory(title="請選擇要分析的資料夾")
    
    if not folder_path:
        print("未選擇資料夾，程式結束")
        sys.exit(0)
        
    return folder_path

def show_completion_message(analysis_result, excel_path, pdf_folder, problematic_count):
    """顯示完成信息對話框"""
    root = tk.Tk()
    root.withdraw()  # 隱藏主窗口
    
    image_files = len(analysis_result["image_files"])
    non_word_files = len(analysis_result["non_word_files"])
    
    message = f"處理完成！\n\n"
    message += f"分析結果:\n"
    message += f"- 發現 {image_files} 個含有圖片的Word文件\n"
    message += f"- 發現 {non_word_files} 個非Word/PDF/TXT檔案\n\n"
    
    if excel_path:
        message += f"Excel報告已生成: {excel_path}\n"
        message += f"共記錄了 {problematic_count} 個問題文件\n\n"
    
    message += f"PDF文件已保存至: {pdf_folder}"
    
    messagebox.showinfo("處理完成", message)

def main():
    print("文件處理整合工具 v1.0")
    print("="*50)
    print("此工具將分析資料夾中的文件，識別問題文件，轉換Word為PDF，並生成Excel報告")
    
    try:
        # 選擇資料夾
        folder_path = select_folder_with_gui()
        
        # 1. 分析資料夾中的文件
        analysis_result = analyze_folder(folder_path)
        
        word_files = [f for f in analysis_result["all_files"] if f.endswith(".docx") or f.endswith(".doc")]
        image_files = analysis_result["image_files"]
        non_word_files = analysis_result["non_word_files"]
        
        # 2. 將Word文件轉換為PDF
        today = datetime.datetime.now().strftime("%Y%m%d")
        pdf_folder = os.path.expanduser(f"~/Documents/會資/Final Project/PDF_2")

        if word_files:
            if sys.platform == 'win32':
                convert_word_to_pdf(word_files, pdf_folder)
            else:
                convert_word_to_pdf_macos(word_files, pdf_folder)
        else:
            print("\n未發現Word文件，跳過PDF轉換步驟")
        
        # 3. 生成Excel報告
        problematic_files = image_files + non_word_files
        if problematic_files:
            excel_path, problematic_count = generate_excel_report(image_files, non_word_files)
        else:
            excel_path, problematic_count = None, 0
            print("\n未發現有問題的文件，不需要生成報告")
        
        # 顯示完成信息
        show_completion_message(analysis_result, excel_path, pdf_folder, problematic_count)
        
        print("\n程式執行完畢！")
    
    except Exception as e:
        print(f"程式執行過程中發生錯誤: {e}")
        messagebox.showerror("錯誤", f"程式執行過程中發生錯誤:\n{e}")
        raise

if __name__ == "__main__":
    main()